from __future__ import annotations

import json
import os
import shutil
import sys
import tempfile
import traceback
from copy import copy
from datetime import date, datetime
from pathlib import Path

FROZEN = bool(getattr(sys, "frozen", False))
BASE_DIR = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))

import webview
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
from openpyxl.utils.datetime import from_excel

APP_NAME = "LTO Vault"
APP_VERSION = "1.0.1"
DEFAULT_SETTINGS = {"sheet_name": "Daily", "tape_column": "A", "date_column": "B", "status_column": "E"}


def data_dir() -> Path:
    folder = Path(os.getenv("LOCALAPPDATA") or Path.home()) / "TapeVault"
    folder.mkdir(parents=True, exist_ok=True)
    return folder


CONFIG = data_dir() / "config.json"
APP_WINDOW = None


def log_error(context: str, error: Exception) -> None:
    try:
        with (data_dir() / "errors.log").open("a", encoding="utf-8") as stream:
            timestamp = datetime.now().isoformat(timespec="seconds")
            stream.write(f"[{timestamp}] {context}: {type(error).__name__}: {error}\n{traceback.format_exc()}\n")
    except OSError:
        pass


def normalized_date(value, epoch) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            result = from_excel(value, epoch)
            return result.date() if isinstance(result, datetime) else result
        except (ValueError, TypeError, OverflowError):
            return None
    if isinstance(value, str):
        for pattern in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value.strip(), pattern).date()
            except ValueError:
                pass
    return None


def tape_label(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def category(status) -> str:
    text = str(status or "").strip().casefold()
    if text in {"tudo certo", "all good", "success", "successful", "ok"}:
        return "success"
    if "fita lotou" in text or "tape full" in text or text.startswith(("falha", "failure", "error")):
        return "danger"
    return "neutral"


class Api:
    def __init__(self) -> None:
        self.path: Path | None = None
        self.custom_tapes: set[str] = set()
        self.settings = dict(DEFAULT_SETTINGS)
        self._cache_key = None
        self._cache_index = None
        self._read_config()

    def _invalidate_cache(self) -> None:
        """Descarta o índice quando a origem dos dados muda."""
        self._cache_key = None
        self._cache_index = None

    def _read_config(self) -> None:
        try:
            saved = json.loads(CONFIG.read_text(encoding="utf-8"))
            candidate = Path(saved.get("workbook", ""))
            self.path = candidate if candidate.is_file() else None
            self.custom_tapes = {tape_label(item) for item in saved.get("custom_tapes", []) if tape_label(item)}
            self.settings.update({key: saved[key] for key in DEFAULT_SETTINGS if saved.get(key)})
        except (OSError, ValueError, TypeError):
            pass

    def _write_config(self) -> None:
        payload = {"workbook": str(self.path or ""), "custom_tapes": sorted(self.custom_tapes), **self.settings}
        CONFIG.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    def get_settings(self):
        return dict(self.settings)

    def save_settings(self, settings):
        cleaned = {
            "sheet_name": str(settings.get("sheet_name", "")).strip(),
            "tape_column": str(settings.get("tape_column", "")).strip().upper(),
            "date_column": str(settings.get("date_column", "")).strip().upper(),
            "status_column": str(settings.get("status_column", "")).strip().upper(),
        }
        if not all(cleaned.values()):
            return {"ok": False, "message": "Preencha todos os campos de mapeamento."}
        try:
            for key in ("tape_column", "date_column", "status_column"):
                column_index_from_string(cleaned[key])
        except ValueError:
            return {"ok": False, "message": "Use letras de colunas do Excel, como A, B ou E."}
        self.settings = cleaned
        self._invalidate_cache()
        self._write_config()
        return {"ok": True}

    def choose_file(self):
        result = APP_WINDOW.create_file_dialog(webview.OPEN_DIALOG, allow_multiple=False, file_types=("Planilhas Excel (*.xlsx;*.xlsm)",))
        if result:
            self.path = Path(result[0]).resolve()
            self._invalidate_cache()
            self._write_config()
            return str(self.path)
        return None

    def get_app_info(self):
        return {"version": APP_VERSION}

    def open_workbook(self):
        if not self.path or not self.path.is_file():
            return {"ok": False, "message": "Selecione uma planilha primeiro."}
        try:
            os.startfile(self.path)
            return {"ok": True}
        except Exception as error:
            log_error("open_workbook", error)
            return {"ok": False, "message": f"Não foi possível abrir a planilha: {error}"}

    def add_tape(self, value: str):
        value = tape_label(value)
        if not value:
            return {"ok": False, "message": "Informe uma identificação válida."}
        self.custom_tapes.add(value)
        self._invalidate_cache()
        self._write_config()
        return {"ok": True, "tape": value}

    def _build_index(self):
        """Lê a planilha uma vez e mantém um índice enquanto o arquivo não mudar."""
        modified = self.path.stat().st_mtime_ns
        cache_key = (str(self.path), modified, tuple(self.settings.values()), tuple(sorted(self.custom_tapes)))
        if cache_key == self._cache_key and self._cache_index is not None:
            return self._cache_index

        sheet_name = self.settings["sheet_name"]
        tape_column = column_index_from_string(self.settings["tape_column"])
        date_column = column_index_from_string(self.settings["date_column"])
        status_column = column_index_from_string(self.settings["status_column"])
        max_column = max(tape_column, date_column, status_column)
        workbook = load_workbook(self.path, read_only=True, data_only=False, keep_vba=self.path.suffix.lower() == ".xlsm", keep_links=True)
        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f'A aba "{sheet_name}" não foi encontrada.')
            sheet = workbook[sheet_name]
            tapes, latest, histories, records, dates = set(self.custom_tapes), {}, {}, [], {}
            for row, values in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_col=max_column, values_only=True), start=2):
                label = tape_label(values[tape_column - 1])
                row_date = normalized_date(values[date_column - 1], workbook.epoch)
                status = values[status_column - 1]
                if label:
                    tapes.add(label)
                if not row_date:
                    continue
                item = {"date": row_date.isoformat(), "tape": label, "status": str(status or "Sem status"), "category": category(status), "row": row}
                records.append(item)
                dates.setdefault(row_date.isoformat(), []).append((row, label, str(status or "")))
                if label:
                    histories.setdefault(label, []).append({key: item[key] for key in ("date", "status", "category", "row")})
                    if status not in (None, "") and (label not in latest or row_date >= latest[label][0]):
                        latest[label] = (row_date, str(status))
        finally:
            workbook.close()

        tape_items = []
        order = lambda item: (0, -int(item)) if item.isdigit() else (1, item.casefold())
        for label in sorted(tapes, key=order):
            last_date, status = latest.get(label, (None, "Sem histórico"))
            entries = sorted(histories.get(label, []), key=lambda item: item["date"], reverse=True)
            tape_items.append({"id": label, "status": status, "category": category(status), "lastDate": last_date.isoformat() if last_date else None, "history": entries})

        self._cache_key = cache_key
        self._cache_index = {"dates": dates, "tapes": tape_items, "records": sorted(records, key=lambda item: (item["date"], item["row"]), reverse=True)}
        return self._cache_index

    def load(self, selected_date: str):
        if not self.path or not self.path.is_file():
            return {"ok": False, "needsFile": True, "message": "Selecione a planilha de backup."}
        try:
            index = self._build_index()
            matches = index["dates"].get(selected_date, [])
            record = matches[0] if len(matches) == 1 else None
            return {
                "ok": True,
                "path": str(self.path),
                "date": selected_date,
                "recordFound": record is not None,
                "duplicate": len(matches) > 1,
                "row": record[0] if record else None,
                "currentTape": record[1] if record else "",
                "currentStatus": record[2] if record else "",
                "tapes": index["tapes"],
                "records": index["records"],
            }
        except Exception as error:
            log_error("load", error)
            return {"ok": False, "message": str(error)}

    def _copy_status_style(self, sheet, row: int, status: str, status_column: int) -> None:
        wanted = category(status)
        target = sheet.cell(row, status_column)
        for previous_row in range(row - 1, 1, -1):
            candidate = sheet.cell(previous_row, status_column)
            if category(candidate.value) == wanted:
                target._style = copy(candidate._style)
                return

    def _copy_tape_style(self, sheet, row: int, tape_column: int) -> None:
        target = sheet.cell(row, tape_column)
        for previous_row in range(row - 1, 1, -1):
            candidate = sheet.cell(previous_row, tape_column)
            if candidate.value not in (None, ""):
                target._style = copy(candidate._style)
                return

    def _append_date_row(self, sheet, target_date: date, date_column: int) -> int:
        row = sheet.max_row + 1
        template_row = row - 1
        if template_row >= 2:
            for column in range(1, sheet.max_column + 1):
                source = sheet.cell(template_row, column)
                destination = sheet.cell(row, column)
                if source.has_style:
                    destination._style = copy(source._style)
            sheet.row_dimensions[row].height = sheet.row_dimensions[template_row].height
        sheet.cell(row, date_column).value = target_date
        return row

    def save(self, selected_date: str, tape: str, status: str):
        if not self.path:
            return {"ok": False, "message": "Nenhuma planilha selecionada."}
        temp_path = None
        workbook = None
        try:
            target_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
            sheet_name = self.settings["sheet_name"]
            tape_column = column_index_from_string(self.settings["tape_column"])
            date_column = column_index_from_string(self.settings["date_column"])
            status_column = column_index_from_string(self.settings["status_column"])
            workbook = load_workbook(self.path, read_only=False, data_only=False, keep_vba=self.path.suffix.lower() == ".xlsm", keep_links=True)
            sheet = workbook[sheet_name]
            # O índice evita percorrer milhares de células novamente ao salvar.
            matches = [item[0] for item in self._build_index()["dates"].get(selected_date, [])]
            if len(matches) > 1:
                raise RuntimeError("A data selecionada aparece mais de uma vez na planilha.")
            created = not matches
            row = matches[0] if matches else self._append_date_row(sheet, target_date, date_column)
            sheet.cell(row, tape_column).value = int(tape) if tape.isdigit() else tape
            self._copy_tape_style(sheet, row, tape_column)
            sheet.cell(row, status_column).value = status
            self._copy_status_style(sheet, row, status, status_column)
            handle, name = tempfile.mkstemp(prefix=f".{self.path.stem}-", suffix=self.path.suffix, dir=self.path.parent)
            os.close(handle)
            temp_path = Path(name)
            workbook.save(temp_path)
            workbook.close()
            workbook = None
            try:
                os.replace(temp_path, self.path)
            except PermissionError:
                shutil.copyfile(temp_path, self.path)
                temp_path.unlink(missing_ok=True)
            temp_path = None
            message = "Novo dia criado e registro salvo com sucesso." if created else "Registro salvo com sucesso."
            self._invalidate_cache()
            return {"ok": True, "message": message}
        except PermissionError as error:
            log_error("save_permission", error)
            return {"ok": False, "message": "A planilha está aberta ou bloqueada. Feche o arquivo no Excel e tente novamente."}
        except Exception as error:
            log_error("save", error)
            return {"ok": False, "message": str(error)}
        finally:
            if workbook:
                workbook.close()
            if temp_path and temp_path.exists():
                temp_path.unlink(missing_ok=True)


if __name__ == "__main__":
    api = Api()
    window = webview.create_window(APP_NAME, str(BASE_DIR / "index.html"), js_api=api, width=1220, height=800, min_size=(980, 680), background_color="#070B14")
    APP_WINDOW = window
    webview.start(gui="edgechromium", debug=False)
