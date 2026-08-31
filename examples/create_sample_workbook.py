from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

output = Path(__file__).with_name("sample-backup-log.xlsx")
workbook = Workbook()
sheet = workbook.active
sheet.title = "Daily"
sheet.append(["TAPE", "DATE", "END", "", "STATUS"])

statuses = ["all good", "all good", "tape full", "all good", "failure: catalog verification"]
start = date.today() - timedelta(days=34)
for index in range(35):
    tape = 61 + (index % 12)
    status = statuses[index % len(statuses)]
    sheet.append([tape, start + timedelta(days=index), None, None, status])

sheet.column_dimensions["A"].width = 12
sheet.column_dimensions["B"].width = 15
sheet.column_dimensions["E"].width = 34
sheet["A1"].font = sheet["B1"].font = sheet["E1"].font = Font(bold=True, color="FFFFFF")
for cell in (sheet["A1"], sheet["B1"], sheet["E1"]):
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center")

for row in range(2, sheet.max_row + 1):
    sheet.cell(row, 1).alignment = Alignment(horizontal="center")
    sheet.cell(row, 2).number_format = "yyyy-mm-dd"
    sheet.cell(row, 5).alignment = Alignment(horizontal="center")

workbook.save(output)
print(output)
